"""
field-story-scorer — Excel field profiling and quality scoring CLI.

Scores every column in an xlsx file across five dimensions (Completeness,
Cardinality, Type Consistency, Distribution, Correlation) and outputs ranked
Excel and PDF reports.

Usage:
    python scorer.py --input data.xlsx --sheet Sheet1 --output-dir ./reports
    python scorer.py --input data.xlsx --sheet Sheet1 --output-dir ./reports --strict-types

See README.md for full documentation.
"""

import argparse
import os
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

__version__ = "1.0.0"


# ---------------------------------------------------------------------------
# Strict-types loader (openpyxl cell-level)
# ---------------------------------------------------------------------------

def load_strict(input_path: str, sheet) -> pd.DataFrame:
    """
    Read an xlsx cell-by-cell via openpyxl, preserving each cell's native Python type
    (int, float, str, bool, datetime, None). Pandas type inference is intentionally
    bypassed so that mixed-type columns are not silently coerced.

    Args:
        input_path: Path to the xlsx file.
        sheet: Sheet name (str) or 0-based index (int).

    Returns:
        DataFrame where every column is dtype=object and values are raw Python types.
    """
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True, read_only=True)

    if isinstance(sheet, int):
        ws = wb.worksheets[sheet]
    else:
        ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return pd.DataFrame()

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data = [list(r) for r in rows[1:]]
    return pd.DataFrame(data, columns=headers)


# ---------------------------------------------------------------------------
# Scoring Engine
# ---------------------------------------------------------------------------

def score_completeness(series: pd.Series) -> float:
    """Ratio of non-null values. 1.0 = fully populated."""
    if len(series) == 0:
        return 0.0
    return round(series.notna().sum() / len(series), 4)


def score_cardinality(series: pd.Series) -> float:
    """
    Normalized uniqueness.
    - Near 0 → constant / near-constant (bad for analytics)
    - Near 1 → every value unique (identifier column)
    Sweet spot is in the middle for categorical fields.
    We return raw ratio; interpretation lives in chart recommendations.
    """
    filled = series.dropna()
    if len(filled) == 0:
        return 0.0
    return round(filled.nunique() / len(filled), 4)


def score_type_consistency(series: pd.Series) -> float:
    """
    Proportion of non-null values that share the majority Python type.

    Standard mode: pandas has already inferred dtypes, so numeric/datetime columns
    score 1.0 by definition. Object columns are inspected via type().

    Strict mode (--strict-types): every column arrives as dtype=object with raw Python
    types preserved. int/float variants are normalized to a single 'numeric' bucket so
    that pandas integer-vs-float coercion differences don't inflate inconsistency.
    """
    filled = series.dropna()
    if len(filled) == 0:
        return 1.0

    # Already typed by pandas — trust it
    if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
        return 1.0

    # Object column: inspect raw Python types.
    # Normalize int/float → 'numeric' so openpyxl int-vs-float noise doesn't penalize
    # columns that are legitimately all-numeric.
    def normalize_type(v):
        t = type(v)
        if t in (int, float):
            return "numeric"
        return t.__name__

    type_counts = filled.map(normalize_type).value_counts()
    majority = type_counts.iloc[0]
    return round(majority / len(filled), 4)


def score_distribution(series: pd.Series) -> float:
    """
    For numeric: normalized coefficient of variation capped at 1.
    For categorical: normalized entropy (0=one category, 1=perfectly uniform).
    Non-numeric / non-categorical gets 0.5 (neutral).
    """
    filled = series.dropna()
    if len(filled) < 2:
        return 0.0

    if pd.api.types.is_numeric_dtype(series):
        mean = filled.mean()
        std = filled.std()
        if mean == 0:
            return 0.0 if std == 0 else 1.0
        cv = abs(std / mean)
        return round(min(cv, 1.0), 4)

    if pd.api.types.is_object_dtype(series) or pd.api.types.is_categorical_dtype(series):
        counts = filled.value_counts(normalize=True)
        k = len(counts)
        if k == 1:
            return 0.0
        entropy = -np.sum(counts * np.log2(counts + 1e-9))
        max_entropy = np.log2(k)
        return round(entropy / max_entropy if max_entropy > 0 else 0.0, 4)

    return 0.5


def score_correlation(series: pd.Series, df: pd.DataFrame) -> float:
    """
    For numeric columns: mean absolute Pearson correlation with all other numeric columns.
    Higher = more correlated to the rest of the dataset (more analytically connected).
    Non-numeric columns get 0.0.
    """
    if not pd.api.types.is_numeric_dtype(series):
        return 0.0
    numeric_df = df.select_dtypes(include=[np.number])
    if numeric_df.shape[1] < 2:
        return 0.0
    try:
        corr_matrix = numeric_df.corr(method="pearson")
        col_name = series.name
        if col_name not in corr_matrix.columns:
            return 0.0
        col_corr = corr_matrix[col_name].drop(col_name, errors="ignore").abs()
        return round(col_corr.mean(), 4)
    except Exception:
        return 0.0


WEIGHTS = {
    "completeness": 0.30,
    "cardinality": 0.15,
    "type_consistency": 0.25,
    "distribution": 0.15,
    "correlation": 0.15,
}


def composite_score(row: dict) -> float:
    total = sum(WEIGHTS[k] * row[k] for k in WEIGHTS)
    return round(total, 4)


def infer_field_type(series: pd.Series) -> str:
    if pd.api.types.is_datetime64_any_dtype(series):
        return "datetime"
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        card = series.dropna().nunique()
        return "numeric_continuous" if card > 20 else "numeric_discrete"
    filled = series.dropna()
    card = filled.nunique()
    n = len(filled)
    if n == 0:
        return "unknown"
    ratio = card / n
    if ratio > 0.9:
        return "identifier"
    if card <= 20:
        return "categorical_low"
    return "categorical_high"


def recommend_chart(field_type: str, cardinality: float, distribution: float) -> str:
    recs = {
        "numeric_continuous": "Histogram, Box Plot, Violin Plot",
        "numeric_discrete": "Bar Chart, Dot Plot",
        "categorical_low": "Bar Chart, Pie Chart (if ≤6 cats), Treemap",
        "categorical_high": "Horizontal Bar Chart (top N), Word Cloud",
        "datetime": "Line Chart, Area Chart, Calendar Heatmap",
        "boolean": "Stacked Bar, Donut Chart",
        "identifier": "Count / Frequency Table — not recommended for visualization",
        "unknown": "N/A — no data",
    }
    base = recs.get(field_type, "Bar Chart")
    if cardinality == 1.0:
        base += " ⚠ All unique — likely ID column"
    elif cardinality < 0.01:
        base += " ⚠ Near-constant — low analytical value"
    return base


# ---------------------------------------------------------------------------
# Analysis Driver
# ---------------------------------------------------------------------------

def analyze(df: pd.DataFrame, strict_types: bool = False) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Returns (rankings_df, profiles_df, chart_recs_df, corr_matrix_df).

    strict_types: when True, df arrived via load_strict() — all columns are dtype=object
    with raw Python types. Correlation is computed by attempting numeric coercion per column
    rather than relying on pandas dtype detection.
    """
    rows = []
    for col in df.columns:
        series = df[col]
        completeness = score_completeness(series)
        cardinality = score_cardinality(series)
        type_consistency = score_type_consistency(series)
        distribution = score_distribution(series)
        correlation = score_correlation(series, df)
        field_type = infer_field_type(series)

        # Build a human-readable type breakdown for strict mode
        if strict_types and pd.api.types.is_object_dtype(series):
            filled = series.dropna()
            def norm(v):
                return "numeric" if isinstance(v, (int, float)) else type(v).__name__
            tc = filled.map(norm).value_counts()
            type_mix = ", ".join(f"{t}:{n}" for t, n in tc.items())
        else:
            type_mix = ""

        dim_scores = {
            "completeness": completeness,
            "cardinality": cardinality,
            "type_consistency": type_consistency,
            "distribution": distribution,
            "correlation": correlation,
        }
        comp = composite_score(dim_scores)

        rows.append({
            "field": col,
            "field_type": field_type,
            "composite_score": comp,
            "completeness": completeness,
            "cardinality": cardinality,
            "type_consistency": type_consistency,
            "distribution": distribution,
            "correlation": correlation,
            "null_count": int(series.isna().sum()),
            "row_count": len(series),
            "unique_count": int(series.nunique()),
            "type_mix": type_mix,
            "chart_recommendation": recommend_chart(field_type, cardinality, distribution),
        })

    all_df = pd.DataFrame(rows)

    rank_cols = ["field", "composite_score", "field_type", "null_count", "unique_count"]
    if strict_types:
        rank_cols.append("type_mix")
    rankings_df = all_df[rank_cols].copy()
    rankings_df = rankings_df.sort_values("composite_score", ascending=False).reset_index(drop=True)
    rankings_df.index += 1
    rankings_df.index.name = "rank"

    profiles_df = all_df[["field", "completeness", "cardinality", "type_consistency",
                            "distribution", "correlation", "composite_score"]].copy()
    profiles_df = profiles_df.sort_values("composite_score", ascending=False).reset_index(drop=True)

    chart_cols = ["field", "field_type", "chart_recommendation", "cardinality", "distribution"]
    if strict_types:
        chart_cols.insert(2, "type_mix")
    chart_recs_df = all_df[chart_cols].copy()
    chart_recs_df = chart_recs_df.sort_values("field").reset_index(drop=True)

    # For strict mode: attempt numeric coercion to build correlation matrix
    if strict_types:
        numeric_cols = {}
        for col in df.columns:
            coerced = pd.to_numeric(df[col], errors="coerce")
            if coerced.notna().sum() > 0 and coerced.notna().mean() > 0.5:
                numeric_cols[col] = coerced
        numeric_df = pd.DataFrame(numeric_cols) if numeric_cols else pd.DataFrame()
    else:
        numeric_df = df.select_dtypes(include=[np.number])

    if not numeric_df.empty and numeric_df.shape[1] >= 2:
        corr_matrix_df = numeric_df.corr(method="pearson").round(4)
    else:
        corr_matrix_df = pd.DataFrame({"note": ["No numeric columns found for correlation"]})

    return rankings_df, profiles_df, chart_recs_df, corr_matrix_df


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

def write_excel(rankings_df, profiles_df, chart_recs_df, corr_matrix_df,
                output_path: str, source_name: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

    wb = Workbook()
    wb.remove(wb.active)

    # Color palette
    HDR_FILL = PatternFill("solid", start_color="1F3864")   # dark navy
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(name="Arial", bold=True, size=13, color="1F3864")
    BODY_FONT = Font(name="Arial", size=10)
    ALT_FILL = PatternFill("solid", start_color="EEF2F7")
    THIN = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_header_row(ws, row_num, col_count, title=None):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def style_data_rows(ws, start_row, end_row, col_count):
        for r in range(start_row, end_row + 1):
            fill = ALT_FILL if r % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            for c in range(1, col_count + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = CENTER if c > 1 else LEFT

    def add_title(ws, title_text, col_span):
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
        t = ws.cell(row=1, column=1, value=title_text)
        t.font = TITLE_FONT
        t.alignment = LEFT
        t.fill = PatternFill("solid", start_color="D9E1F2")

    def auto_width(ws, min_w=10, max_w=40):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)

    def df_to_sheet(ws, df, include_index=False):
        if include_index:
            df = df.reset_index()
        headers = list(df.columns)
        ws.append(headers)
        for _, row in df.iterrows():
            ws.append([str(v) if not isinstance(v, (int, float)) else v for v in row])
        return len(headers)

    # --- Tab 1: Field Rankings ---
    ws1 = wb.create_sheet("Field Rankings")
    col_count = df_to_sheet(ws1, rankings_df, include_index=True)
    add_title(ws1, f"Field Rankings — {source_name}", col_count)
    style_header_row(ws1, 2, col_count)
    style_data_rows(ws1, 3, ws1.max_row, col_count)
    auto_width(ws1)

    # Color scale on composite score column (col 3 after rank + field)
    score_col = get_column_letter(3)
    data_start = f"{score_col}3"
    data_end = f"{score_col}{ws1.max_row}"
    ws1.conditional_formatting.add(
        f"{data_start}:{data_end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="num", mid_value=0.5, mid_color="FFEB84",
                       end_type="num", end_value=1, end_color="63BE7B")
    )
    ws1.freeze_panes = "A3"

    # --- Tab 2: Field Profiles ---
    ws2 = wb.create_sheet("Field Profiles")
    col_count2 = df_to_sheet(ws2, profiles_df)
    add_title(ws2, f"Field Profiles — Dimension Breakdown — {source_name}", col_count2)
    style_header_row(ws2, 2, col_count2)
    style_data_rows(ws2, 3, ws2.max_row, col_count2)
    auto_width(ws2)

    # Data bars on each score column (cols 2–7)
    for c_idx in range(2, col_count2 + 1):
        col_letter = get_column_letter(c_idx)
        ws2.conditional_formatting.add(
            f"{col_letter}3:{col_letter}{ws2.max_row}",
            DataBarRule(start_type="num", start_value=0, end_type="num",
                        end_value=1, color="4472C4")
        )
    ws2.freeze_panes = "B3"

    # --- Tab 3: Chart Recommendations ---
    ws3 = wb.create_sheet("Chart Recommendations")
    col_count3 = df_to_sheet(ws3, chart_recs_df)
    add_title(ws3, f"Chart Recommendations — {source_name}", col_count3)
    style_header_row(ws3, 2, col_count3)
    style_data_rows(ws3, 3, ws3.max_row, col_count3)
    auto_width(ws3)
    ws3.freeze_panes = "A3"

    # --- Tab 4: Correlation Matrix ---
    ws4 = wb.create_sheet("Correlation Matrix")
    if "note" in corr_matrix_df.columns:
        ws4.append(["note"])
        ws4.append([corr_matrix_df["note"].iloc[0]])
    else:
        col_count4 = df_to_sheet(ws4, corr_matrix_df.reset_index())
        add_title(ws4, f"Pearson Correlation Matrix (Numeric Fields) — {source_name}", col_count4)
        style_header_row(ws4, 2, col_count4)
        style_data_rows(ws4, 3, ws4.max_row, col_count4)
        auto_width(ws4)
        # Diverging color scale
        score_col_end = get_column_letter(col_count4)
        ws4.conditional_formatting.add(
            f"B3:{score_col_end}{ws4.max_row}",
            ColorScaleRule(start_type="num", start_value=-1, start_color="F8696B",
                           mid_type="num", mid_value=0, mid_color="FFFFFF",
                           end_type="num", end_value=1, end_color="63BE7B")
        )
        ws4.freeze_panes = "B3"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# PDF Report
# ---------------------------------------------------------------------------

def write_pdf(rankings_df, profiles_df, chart_recs_df, corr_matrix_df,
              output_path: str, source_name: str):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                     TableStyle, PageBreak, HRFlowable)
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    PAGE_W, PAGE_H = landscape(letter)
    NAVY = colors.HexColor("#1F3864")
    LIGHT_BLUE = colors.HexColor("#D9E1F2")
    ALT_ROW = colors.HexColor("#EEF2F7")
    GREEN = colors.HexColor("#63BE7B")
    RED = colors.HexColor("#F8696B")
    YELLOW = colors.HexColor("#FFEB84")

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16,
                         textColor=NAVY, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12,
                         textColor=NAVY, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, spaceAfter=2)
    caption = ParagraphStyle("caption", parent=styles["Normal"], fontSize=8,
                              textColor=colors.grey, spaceAfter=6)

    def score_color(val):
        try:
            v = float(val)
            if v >= 0.75:
                return GREEN
            if v >= 0.45:
                return YELLOW
            return RED
        except:
            return colors.white

    def make_table(df, col_widths=None, score_cols=None):
        data = [list(df.columns)]
        for _, row in df.iterrows():
            data.append([str(v) if not isinstance(v, (int, float)) else round(v, 4) for v in row])

        available_w = PAGE_W - inch
        if col_widths is None:
            w = available_w / len(df.columns)
            col_widths = [w] * len(df.columns)

        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ALT_ROW]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        if score_cols:
            for r_idx, row in enumerate(data[1:], start=1):
                for c_idx in score_cols:
                    if c_idx < len(row):
                        cell_val = row[c_idx]
                        bg = score_color(cell_val)
                        style_cmds.append(("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), bg))

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle(style_cmds))
        return t

    story = []

    # Cover-style header
    story.append(Paragraph(f"Field Story Scorer Report", h1))
    story.append(Paragraph(f"Source: {source_name}", body))
    story.append(HRFlowable(width="100%", thickness=2, color=NAVY))
    story.append(Spacer(1, 0.15 * inch))

    # Tab 1: Rankings
    story.append(Paragraph("Tab 1 — Field Rankings", h2))
    story.append(Paragraph(
        "Fields ranked by composite score (weighted: Completeness 30%, Type Consistency 25%, "
        "Cardinality 15%, Distribution 15%, Correlation 15%). "
        "Green ≥ 0.75 | Yellow 0.45–0.74 | Red < 0.45.", caption))

    r_df = rankings_df.reset_index()
    n_cols = len(r_df.columns)
    available = PAGE_W - inch
    col_widths_r = [0.5 * inch, 2.2 * inch] + [(available - 2.7 * inch) / (n_cols - 2)] * (n_cols - 2)
    story.append(make_table(r_df, col_widths_r, score_cols=[2]))
    story.append(PageBreak())

    # Tab 2: Profiles
    story.append(Paragraph("Tab 2 — Field Profiles (Dimension Breakdown)", h2))
    story.append(Paragraph(
        "Per-field scores across all five dimensions. Scores are 0–1 ratios. "
        "Color coding matches the ranking table.", caption))
    p_df = profiles_df.copy()
    n_cols_p = len(p_df.columns)
    first_col_w = 2.0 * inch
    rest_w = (PAGE_W - inch - first_col_w) / (n_cols_p - 1)
    col_widths_p = [first_col_w] + [rest_w] * (n_cols_p - 1)
    score_cols_p = list(range(1, n_cols_p))
    story.append(make_table(p_df, col_widths_p, score_cols=score_cols_p))
    story.append(PageBreak())

    # Tab 3: Chart Recommendations
    story.append(Paragraph("Tab 3 — Chart Recommendations", h2))
    story.append(Paragraph(
        "Suggested visualization types per field, derived from inferred field type, "
        "cardinality ratio, and distribution score.", caption))
    c_df = chart_recs_df.copy()
    col_widths_c = [1.8 * inch, 1.5 * inch, 4.5 * inch, 1.0 * inch, 1.0 * inch]
    total_c = sum(col_widths_c)
    if total_c > PAGE_W - inch:
        scale = (PAGE_W - inch) / total_c
        col_widths_c = [w * scale for w in col_widths_c]
    story.append(make_table(c_df, col_widths_c))
    story.append(PageBreak())

    # Tab 4: Correlation Matrix
    story.append(Paragraph("Tab 4 — Pearson Correlation Matrix (Numeric Fields)", h2))
    if "note" in corr_matrix_df.columns:
        story.append(Paragraph(corr_matrix_df["note"].iloc[0], body))
    else:
        story.append(Paragraph(
            "Pearson correlation coefficients for all numeric fields. "
            "Green = strong positive, Red = strong negative, White = near-zero.", caption))
        cm = corr_matrix_df.reset_index()
        n_c = len(cm.columns)
        cw = (PAGE_W - inch) / n_c
        col_widths_cm = [cw] * n_c
        score_cols_cm = list(range(1, n_c))
        story.append(make_table(cm, col_widths_cm, score_cols=score_cols_cm))

    doc.build(story)


# ---------------------------------------------------------------------------
# CLI Entry Point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="field-story-scorer: Score every column in an xlsx file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="See README.md for full documentation and scoring methodology.",
    )
    parser.add_argument("--version", action="version", version=f"field-story-scorer {__version__}")
    parser.add_argument("--input", required=True, help="Path to input Excel file")
    parser.add_argument("--sheet", default=0, help="Sheet name or index (default: first sheet)")
    parser.add_argument("--output-dir", default="./reports", help="Directory for output files")
    parser.add_argument(
        "--strict-types",
        action="store_true",
        help=(
            "Read cell values directly via openpyxl instead of letting pandas infer dtypes. "
            "Use this when columns contain mixed types that pandas silently coerces — e.g. a "
            "column that is mostly numeric but contains stray strings like 'N/A', 'TBD', or '—'. "
            "Without this flag, pandas converts those strings to NaN and the column scores as "
            "fully numeric. With this flag, every cell's raw Python type is preserved and "
            "type_consistency scores reflect true cell-level heterogeneity. A 'type_mix' column "
            "is added to the rankings and chart-recs tabs showing the exact type breakdown."
        ),
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    sheet_arg = args.sheet
    try:
        sheet_arg = int(sheet_arg)
    except ValueError:
        pass

    if args.strict_types:
        print(f"Reading (strict-types): {input_path} | Sheet: {sheet_arg}")
        try:
            df = load_strict(str(input_path), sheet_arg)
        except Exception as e:
            print(f"ERROR reading file (strict mode): {e}", file=sys.stderr)
            sys.exit(1)
    else:
        print(f"Reading: {input_path} | Sheet: {sheet_arg}")
        try:
            df = pd.read_excel(input_path, sheet_name=sheet_arg)
        except Exception as e:
            print(f"ERROR reading file: {e}", file=sys.stderr)
            sys.exit(1)

    print(f"  {len(df)} rows × {len(df.columns)} columns")
    if args.strict_types:
        print("  [strict-types] pandas dtype inference bypassed — raw cell types preserved")

    print("Scoring fields...")
    rankings_df, profiles_df, chart_recs_df, corr_matrix_df = analyze(df, strict_types=args.strict_types)

    stem = input_path.stem
    mode_tag = "_strict" if args.strict_types else ""
    source_name = f"{stem} ({sheet_arg}){' [strict-types]' if args.strict_types else ''}"

    xlsx_out = output_dir / f"{stem}_field_report{mode_tag}.xlsx"
    pdf_out = output_dir / f"{stem}_field_report{mode_tag}.pdf"

    print(f"Writing Excel → {xlsx_out}")
    write_excel(rankings_df, profiles_df, chart_recs_df, corr_matrix_df,
                str(xlsx_out), source_name)

    print(f"Writing PDF  → {pdf_out}")
    write_pdf(rankings_df, profiles_df, chart_recs_df, corr_matrix_df,
              str(pdf_out), source_name)

    print("\nDone.")
    print(f"  Excel: {xlsx_out}")
    print(f"  PDF:   {pdf_out}")

    print("\nTop 5 fields by composite score:")
    top5 = rankings_df.head(5).reset_index()
    for _, row in top5.iterrows():
        mix = f"  mix=[{row['type_mix']}]" if args.strict_types and row.get("type_mix") else ""
        print(f"  #{row['rank']:>2}  {row['field']:<30}  score={row['composite_score']:.4f}  type={row['field_type']}{mix}")


if __name__ == "__main__":
    main()

